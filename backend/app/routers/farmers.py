import io
from typing import Optional, List
from datetime import date
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from ..database import get_pool

router = APIRouter()


# ---------- Pydantic schemas ----------

class FarmerCreate(BaseModel):
    full_name: str
    iin: str
    contract_number: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None


class FarmerUpdate(BaseModel):
    full_name: Optional[str] = None
    iin: Optional[str] = None
    contract_number: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None


class FarmerResponse(BaseModel):
    id: int
    full_name: str
    iin: str
    contract_number: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    created_at: str
    updated_at: str


# ---------- CRUD endpoints ----------

@router.get("/farmers", response_model=List[FarmerResponse])
async def list_farmers(search: Optional[str] = Query(None, description="Поиск по ФИО, ИИН или номеру договора")):
    pool = get_pool()
    if search:
        term = f"%{search.strip()}%"
        rows = await pool.fetch("""
            SELECT id, full_name, iin, contract_number, phone, address, created_at, updated_at
            FROM farmers
            WHERE full_name ILIKE $1 OR iin ILIKE $1 OR contract_number ILIKE $1
            ORDER BY id
        """, term)
    else:
        rows = await pool.fetch("""
            SELECT id, full_name, iin, contract_number, phone, address, created_at, updated_at
            FROM farmers ORDER BY id
        """)
    return [
        {
            "id": r["id"], "full_name": r["full_name"], "iin": r["iin"],
            "contract_number": r["contract_number"], "phone": r["phone"],
            "address": r["address"],
            "created_at": r["created_at"].isoformat(),
            "updated_at": r["updated_at"].isoformat(),
        }
        for r in rows
    ]


@router.get("/farmers/{farmer_id}", response_model=FarmerResponse)
async def get_farmer(farmer_id: int):
    pool = get_pool()
    row = await pool.fetchrow("""
        SELECT id, full_name, iin, contract_number, phone, address, created_at, updated_at
        FROM farmers WHERE id = $1
    """, farmer_id)
    if not row:
        raise HTTPException(status_code=404, detail="Крестьянин не найден")
    return {
        "id": row["id"], "full_name": row["full_name"], "iin": row["iin"],
        "contract_number": row["contract_number"], "phone": row["phone"],
        "address": row["address"],
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }


@router.post("/farmers", response_model=FarmerResponse, status_code=201)
async def create_farmer(body: FarmerCreate):
    pool = get_pool()
    try:
        row = await pool.fetchrow("""
            INSERT INTO farmers (full_name, iin, contract_number, phone, address)
            VALUES ($1, $2, $3, $4, $5)
            RETURNING id, full_name, iin, contract_number, phone, address, created_at, updated_at
        """, body.full_name, body.iin, body.contract_number, body.phone, body.address)
    except Exception as e:
        err = str(e)
        if "unique" in err.lower() or "duplicate" in err.lower():
            raise HTTPException(status_code=409, detail="Крестьянин с таким ИИН уже существует")
        raise HTTPException(status_code=500, detail=err)
    return {
        "id": row["id"], "full_name": row["full_name"], "iin": row["iin"],
        "contract_number": row["contract_number"], "phone": row["phone"],
        "address": row["address"],
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }


@router.put("/farmers/{farmer_id}", response_model=FarmerResponse)
async def update_farmer(farmer_id: int, body: FarmerUpdate):
    pool = get_pool()
    existing = await pool.fetchrow("SELECT id FROM farmers WHERE id = $1", farmer_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Крестьянин не найден")

    fields = []
    values = []
    idx = 1
    for col in ["full_name", "iin", "contract_number", "phone", "address"]:
        val = getattr(body, col, None)
        if val is not None:
            fields.append(f"{col} = ${idx}")
            values.append(val)
            idx += 1

    if not fields:
        raise HTTPException(status_code=400, detail="Нет полей для обновления")

    fields.append("updated_at = NOW()")
    values.append(farmer_id)

    try:
        row = await pool.fetchrow(f"""
            UPDATE farmers SET {', '.join(fields)} WHERE id = ${idx}
            RETURNING id, full_name, iin, contract_number, phone, address, created_at, updated_at
        """, *values)
    except Exception as e:
        err = str(e)
        if "unique" in err.lower() or "duplicate" in err.lower():
            raise HTTPException(status_code=409, detail="Крестьянин с таким ИИН уже существует")
        raise HTTPException(status_code=500, detail=err)

    return {
        "id": row["id"], "full_name": row["full_name"], "iin": row["iin"],
        "contract_number": row["contract_number"], "phone": row["phone"],
        "address": row["address"],
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }


@router.delete("/farmers/{farmer_id}")
async def delete_farmer(farmer_id: int):
    pool = get_pool()
    row = await pool.fetchrow("DELETE FROM farmers WHERE id = $1 RETURNING id", farmer_id)
    if not row:
        raise HTTPException(status_code=404, detail="Крестьянин не найден")
    return {"deleted": True}


# ---------- Excel import endpoint ----------

@router.post("/farmers/import-excel")
async def import_farmers_excel(file: UploadFile = File(...)):
    """
    Import farmers from Excel (.xlsx / .xls).
    Expected columns: full_name*, iin*, contract_number, phone, address
    """
    if not file.filename or not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Поддерживаются только .xlsx / .xls файлы")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен на сервере")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip().lower().replace(" ", "_"))
        else:
            headers.append("")

    required = ["full_name", "iin"]
    for req in required:
        if req not in headers:
            raise HTTPException(status_code=400, detail=f"В Excel отсутствует колонка: {req}")

    def col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    idx_name = col("full_name")
    idx_iin = col("iin")
    idx_contract_number = col("contract_number")
    idx_phone = col("phone")
    idx_address = col("address")

    pool = get_pool()
    imported = 0
    errors = []

    async with pool.acquire() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            full_name = str(row[idx_name]).strip() if idx_name >= 0 and row[idx_name] else None
            iin = str(row[idx_iin]).strip() if idx_iin >= 0 and row[idx_iin] else None
            if not full_name or not iin:
                continue

            contract_number = str(row[idx_contract_number]).strip() if idx_contract_number >= 0 and row[idx_contract_number] else None
            phone = str(row[idx_phone]).strip() if idx_phone >= 0 and row[idx_phone] else None
            address = str(row[idx_address]).strip() if idx_address >= 0 and row[idx_address] else None

            try:
                async with conn.transaction():
                    existing = await conn.fetchrow("SELECT id FROM farmers WHERE iin = $1", iin)
                    if existing:
                        await conn.execute(
                            "UPDATE farmers SET full_name=$1, contract_number=$2, phone=$3, address=$4 WHERE id=$5",
                            full_name, contract_number, phone, address, existing["id"],
                        )
                    else:
                        await conn.execute(
                            """INSERT INTO farmers (full_name, iin, contract_number, phone, address)
                               VALUES ($1, $2, $3, $4, $5)""",
                            full_name, iin, contract_number, phone, address,
                        )
                        imported += 1
            except Exception as e:
                errors.append(f"Строка ({full_name}, {iin}): {e}")

    return {
        "imported": imported,
        "errors": errors,
    }
