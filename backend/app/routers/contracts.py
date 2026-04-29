import io
import json
from typing import Optional, List
from datetime import date
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from ..database import get_pool

router = APIRouter()


# ---------- Pydantic schemas ----------

class ContractParcelItem(BaseModel):
    """One parcel within a contract (input)."""
    parcel_id: Optional[int] = None
    cadastral_number: Optional[str] = None
    distribution_channel: Optional[str] = None
    main_channel: Optional[str] = None
    culture: Optional[str] = None
    doc_hectares: Optional[float] = None
    irrigated_hectares: Optional[float] = None
    rural_district: Optional[str] = None
    # GeoJSON geometry as dict (optional, set later via coordinates)
    geom: Optional[dict] = None


class ContractParcelResponse(ContractParcelItem):
    id: int
    contract_id: int
    created_at: str
    updated_at: str


class ContractCreate(BaseModel):
    farmer_id: int
    contract_number: str
    contract_date: Optional[date] = None
    total_water_volume: Optional[float] = None
    actual_water_volume: Optional[float] = None
    tariff_amount: Optional[float] = None
    year: Optional[int] = None
    parcels: List[ContractParcelItem] = []


class ContractUpdate(BaseModel):
    contract_number: Optional[str] = None
    contract_date: Optional[date] = None
    total_water_volume: Optional[float] = None
    actual_water_volume: Optional[float] = None
    tariff_amount: Optional[float] = None
    year: Optional[int] = None


class ContractResponse(BaseModel):
    id: int
    farmer_id: int
    farmer_name: Optional[str] = None
    contract_number: str
    contract_date: Optional[str] = None
    total_water_volume: Optional[float] = None
    actual_water_volume: Optional[float] = None
    tariff_amount: Optional[float] = None
    year: Optional[int] = None
    created_at: str
    updated_at: str
    parcels: List[ContractParcelResponse] = []


# ---------- Tariff defaults ----------
# 2019-2024 → 1.0 тг, 2026 → 1.8 тг
TARIFF_BY_YEAR = {
    2019: 1.0, 2020: 1.0, 2021: 1.0, 2022: 1.0, 2023: 1.0, 2024: 1.0,
    2026: 1.8,
}


def default_tariff(year: Optional[int]) -> Optional[float]:
    if year is None:
        return None
    return TARIFF_BY_YEAR.get(year)


# ---------- Helpers ----------

async def _fetch_contract_parcels(pool, contract_id: int) -> List[dict]:
    rows = await pool.fetch("""
        SELECT cp.id, cp.contract_id, cp.parcel_id, cp.cadastral_number,
               cp.distribution_channel, cp.main_channel, cp.culture,
               cp.doc_hectares, cp.irrigated_hectares, cp.rural_district,
               ST_AsGeoJSON(COALESCE(cp.geom, p.geom))::jsonb AS geom_json,
               cp.created_at, cp.updated_at
        FROM contract_parcels cp
        LEFT JOIN parcels p ON p.id = cp.parcel_id
        WHERE cp.contract_id = $1 ORDER BY cp.id
    """, contract_id)
    return [
        {
            "id": r["id"], "contract_id": r["contract_id"],
            "parcel_id": r["parcel_id"],
            "cadastral_number": r["cadastral_number"],
            "distribution_channel": r["distribution_channel"],
            "main_channel": r["main_channel"],
            "culture": r["culture"],
            "doc_hectares": float(r["doc_hectares"]) if r["doc_hectares"] is not None else None,
            "irrigated_hectares": float(r["irrigated_hectares"]) if r["irrigated_hectares"] is not None else None,
            "rural_district": r["rural_district"],
            "geom": json.loads(r["geom_json"]) if r["geom_json"] else None,
            "created_at": r["created_at"].isoformat(),
            "updated_at": r["updated_at"].isoformat(),
        }
        for r in rows
    ]


async def _build_contract_response(row, pool) -> dict:
    parcels = await _fetch_contract_parcels(pool, row["id"])
    return {
        "id": row["id"], "farmer_id": row["farmer_id"],
        "farmer_name": row["farmer_name"],
        "contract_number": row["contract_number"],
        "contract_date": row["contract_date"].isoformat() if row["contract_date"] else None,
        "total_water_volume": float(row["total_water_volume"]) if row["total_water_volume"] is not None else None,
        "actual_water_volume": float(row["actual_water_volume"]) if row["actual_water_volume"] is not None else None,
        "tariff_amount": float(row["tariff_amount"]) if row["tariff_amount"] is not None else None,
        "year": row["year"],
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
        "parcels": parcels,
    }


CONTRACT_WITH_FARMER = """
    SELECT c.id, c.farmer_id, f.full_name AS farmer_name,
           c.contract_number, c.contract_date, c.total_water_volume,
           c.actual_water_volume, c.tariff_amount,
           c.year, c.created_at, c.updated_at
    FROM contracts c
    LEFT JOIN farmers f ON f.id = c.farmer_id
"""


# ---------- CRUD endpoints ----------

@router.get("/contracts", response_model=List[ContractResponse])
async def list_contracts(
    farmer_id: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    search: Optional[str] = Query(None, description="Поиск по номеру договора"),
):
    pool = get_pool()
    conditions = []
    params = []
    idx = 1

    if farmer_id is not None:
        conditions.append(f"c.farmer_id = ${idx}")
        params.append(farmer_id)
        idx += 1
    if year is not None:
        conditions.append(f"c.year = ${idx}")
        params.append(year)
        idx += 1
    if search:
        conditions.append(f"c.contract_number ILIKE ${idx}")
        params.append(f"%{search.strip()}%")
        idx += 1

    where = ""
    if conditions:
        where = " WHERE " + " AND ".join(conditions)

    rows = await pool.fetch(CONTRACT_WITH_FARMER + where + " ORDER BY c.year DESC, c.id", *params)

    result = []
    for r in rows:
        result.append(await _build_contract_response(r, pool))
    return result


# ---------- Analytics / Stats endpoints (срезки) ----------

@router.get("/contracts/stats")
async def get_contracts_stats(
    year: Optional[int] = Query(None),
    rural_district: Optional[str] = Query(None),
    culture: Optional[str] = Query(None),
    farmer_id: Optional[int] = Query(None),
):
    """
    Aggregated statistics on contracts and parcels.
    Returns totals by volume, hectares, tariff — filterable by year, district, culture, farmer.
    """
    pool = get_pool()

    # Build shared where clauses for both tables
    contract_conds = []
    parcel_conds = []
    params = []
    idx = 1

    if year is not None:
        contract_conds.append(f"c.year = ${idx}")
        parcel_conds.append(f"c.year = ${idx}")
        params.append(year)
        idx += 1
    if farmer_id is not None:
        contract_conds.append(f"c.farmer_id = ${idx}")
        parcel_conds.append(f"c.farmer_id = ${idx}")
        params.append(farmer_id)
        idx += 1
    if rural_district:
        parcel_conds.append(f"cp.rural_district ILIKE ${idx}")
        params.append(f"%{rural_district.strip()}%")
        idx += 1
    if culture:
        parcel_conds.append(f"cp.culture ILIKE ${idx}")
        params.append(f"%{culture.strip()}%")
        idx += 1

    c_where = ""
    if contract_conds:
        c_where = " WHERE " + " AND ".join(contract_conds)
    p_where = ""
    if parcel_conds:
        p_where = " WHERE " + " AND ".join(parcel_conds)

    # Contract-level aggregation
    c_row = await pool.fetchrow(f"""
        SELECT
            COUNT(*) AS total_contracts,
            COALESCE(SUM(c.total_water_volume), 0) AS total_plan_volume,
            COALESCE(SUM(c.actual_water_volume), 0) AS total_actual_volume,
            COALESCE(SUM(c.tariff_amount), 0) AS total_tariff
        FROM contracts c
        {c_where}
    """, *params)

    # Parcel-level aggregation
    p_row = await pool.fetchrow(f"""
        SELECT
            COUNT(*) AS total_parcels,
            COALESCE(SUM(cp.doc_hectares), 0) AS total_doc_hectares,
            COALESCE(SUM(cp.irrigated_hectares), 0) AS total_irrigated_hectares
        FROM contract_parcels cp
        JOIN contracts c ON c.id = cp.contract_id
        {p_where}
    """, *params)

    # By year breakdown
    years_rows = await pool.fetch(f"""
        SELECT c.year,
               COUNT(*) AS cnt,
               COALESCE(SUM(c.total_water_volume), 0) AS plan_vol,
               COALESCE(SUM(c.actual_water_volume), 0) AS actual_vol,
               COALESCE(SUM(c.tariff_amount), 0) AS tariff
        FROM contracts c
        {c_where}
        GROUP BY c.year ORDER BY c.year
    """, *params)

    # By culture breakdown
    culture_rows = await pool.fetch(f"""
        SELECT cp.culture,
               COUNT(*) AS cnt,
               COALESCE(SUM(cp.doc_hectares), 0) AS doc_ha,
               COALESCE(SUM(cp.irrigated_hectares), 0) AS irr_ha
        FROM contract_parcels cp
        JOIN contracts c ON c.id = cp.contract_id
        {p_where}
        GROUP BY cp.culture ORDER BY cnt DESC
    """, *params)

    return {
        "overview": {
            "total_contracts": c_row["total_contracts"],
            "total_plan_volume": float(c_row["total_plan_volume"]),
            "total_actual_volume": float(c_row["total_actual_volume"]),
            "total_tariff": float(c_row["total_tariff"]),
            "total_parcels": p_row["total_parcels"],
            "total_doc_hectares": float(p_row["total_doc_hectares"]),
            "total_irrigated_hectares": float(p_row["total_irrigated_hectares"]),
        },
        "by_year": [
            {
                "year": r["year"],
                "contracts": r["cnt"],
                "plan_volume": float(r["plan_vol"]),
                "actual_volume": float(r["actual_vol"]),
                "tariff": float(r["tariff"]),
            }
            for r in years_rows
        ],
        "by_culture": [
            {
                "culture": r["culture"],
                "parcels": r["cnt"],
                "doc_hectares": float(r["doc_ha"]),
                "irrigated_hectares": float(r["irr_ha"]),
            }
            for r in culture_rows
        ],
    }


# ---------- Colour helpers (same as parcels router) ----------

SRID = 32642


def cp_fill_color(pct: float) -> list[int]:
    if pct < 60:
        return [52, 199, 89, 170]
    if pct < 80:
        return [245, 158, 11, 175]
    if pct < 95:
        return [249, 115, 22, 185]
    return [255, 59, 48, 195]


def cp_status_label(pct: float) -> str:
    if pct < 60:
        return "Норма"
    if pct < 80:
        return "Внимание"
    if pct < 95:
        return "Высокое"
    return "Критично"


# ---------- GeoJSON endpoint for contract_parcels on map ----------

def build_cp_map_query(where_clause: str) -> str:
    """Build contract parcels map query with dynamic WHERE clause (string concatenation)."""
    srid = SRID
    query = (
        "WITH cp_bounds AS ("
        "    SELECT cp.id, cp.contract_id, cp.parcel_id, cp.cadastral_number,"
        "           cp.distribution_channel, cp.main_channel, cp.culture,"
        "           cp.doc_hectares, cp.irrigated_hectares, cp.rural_district,"
        "           COALESCE(cp.geom, p.geom) AS geom_native,"
        "           ST_Transform(COALESCE(cp.geom, p.geom), " + str(srid) + ") AS geom_utm,"
        "           CASE WHEN cp.doc_hectares > 0 AND cp.irrigated_hectares IS NOT NULL"
        "                THEN LEAST(ROUND((cp.irrigated_hectares / cp.doc_hectares * 100)::NUMERIC, 2), 100)"
        "                ELSE 0 END AS fill_pct,"
        "           ST_XMin(ST_Transform(COALESCE(cp.geom, p.geom), " + str(srid) + ")) AS xmin,"
        "           ST_YMin(ST_Transform(COALESCE(cp.geom, p.geom), " + str(srid) + ")) AS ymin,"
        "           ST_XMax(ST_Transform(COALESCE(cp.geom, p.geom), " + str(srid) + ")) AS xmax,"
        "           ST_YMax(ST_Transform(COALESCE(cp.geom, p.geom), " + str(srid) + ")) AS ymax,"
        "           c.year, c.contract_number, f.full_name AS farmer_name"
        "    FROM contract_parcels cp"
        "    JOIN contracts c ON c.id = cp.contract_id"
        "    JOIN farmers f ON f.id = c.farmer_id"
        "    LEFT JOIN parcels p ON p.id = cp.parcel_id"
        "    WHERE COALESCE(cp.geom, p.geom) IS NOT NULL"
        "        " + where_clause +
        "),"
        "cp_filled AS ("
        "    SELECT id, contract_id, parcel_id, cadastral_number,"
        "           distribution_channel, main_channel, culture,"
        "           doc_hectares, irrigated_hectares, rural_district,"
        "           fill_pct, year, contract_number, farmer_name,"
        "           geom_utm AS full_geom,"
        "           CASE"
        "               WHEN fill_pct >= 100 THEN geom_utm"
        "               WHEN fill_pct <= 0   THEN NULL"
        "               ELSE ST_CollectionExtract("
        "                   ST_Intersection("
        "                       geom_utm,"
        "                       ST_SetSRID(ST_MakeBox2D("
        "                           ST_Point(xmin - 1, ymin - 1),"
        "                           ST_Point(xmax + 1, ymin + (ymax - ymin) * fill_pct / 100.0)"
        "                       ), " + str(srid) + ")"
        "                   ), 3)"
        "           END AS clipped_geom"
        "    FROM cp_bounds"
        ")"
        "SELECT id, contract_id, parcel_id, cadastral_number,"
        "       distribution_channel, main_channel, culture,"
        "       doc_hectares, irrigated_hectares, rural_district,"
        "       fill_pct, year, contract_number, farmer_name,"
        "       ST_AsGeoJSON(ST_Transform(full_geom, 4326))::jsonb AS outline_geom,"
        "       ST_AsGeoJSON(ST_Transform(clipped_geom, 4326))::jsonb AS fill_geom,"
        "       ST_AsGeoJSON(ST_Centroid(ST_Transform(full_geom, 4326)))::jsonb AS centroid_geom"
        "FROM cp_filled ORDER BY id"
    )
    return query


@router.get("/contracts/parcels/map")
async def get_contract_parcels_geojson(
    year: Optional[int] = Query(None),
    rural_district: Optional[str] = Query(None),
    culture: Optional[str] = Query(None),
):
    """Return contract parcels with outlines + dynamic fills (like /parcels/map)."""
    pool = get_pool()
    conditions = []
    params = []
    idx = 1

    if year is not None:
        conditions.append(f"c.year = ${idx}")
        params.append(year)
        idx += 1
    if rural_district:
        conditions.append(f"cp.rural_district ILIKE ${idx}")
        params.append(f"%{rural_district.strip()}%")
        idx += 1
    if culture:
        conditions.append(f"cp.culture ILIKE ${idx}")
        params.append(f"%{culture.strip()}%")
        idx += 1

    where_clause = ""
    if conditions:
        where_clause = "AND " + " AND ".join(conditions)

    query = build_cp_map_query(where_clause)
    rows = await pool.fetch(query, *params)

    outlines, fills = [], []
    for row in rows:
        pct = float(row["fill_pct"])
        centroid = row["centroid_geom"]
        lon = centroid["coordinates"][0] if centroid else None
        lat = centroid["coordinates"][1] if centroid else None
        props = {
            "id": row["id"],
            "contract_id": row["contract_id"],
            "parcel_id": row["parcel_id"],
            "cadastral_number": row["cadastral_number"],
            "distribution_channel": row["distribution_channel"],
            "main_channel": row["main_channel"],
            "culture": row["culture"],
            "doc_hectares": float(row["doc_hectares"]) if row["doc_hectares"] is not None else None,
            "irrigated_hectares": float(row["irrigated_hectares"]) if row["irrigated_hectares"] is not None else None,
            "rural_district": row["rural_district"],
            "year": row["year"],
            "contract_number": row["contract_number"],
            "farmer_name": row["farmer_name"],
            "fill_pct": pct,
            "status": cp_status_label(pct),
            "fill_color": cp_fill_color(pct),
            "lon": lon,
            "lat": lat,
        }
        outlines.append({"type": "Feature", "geometry": row["outline_geom"], "properties": props})
        if row["fill_geom"]:
            fills.append({"type": "Feature", "geometry": row["fill_geom"], "properties": props})

    return {
        "outlines": {"type": "FeatureCollection", "features": outlines},
        "fills": {"type": "FeatureCollection", "features": fills},
    }


@router.get("/contracts/{contract_id}", response_model=ContractResponse)
async def get_contract(contract_id: int):
    pool = get_pool()
    row = await pool.fetchrow(CONTRACT_WITH_FARMER + " WHERE c.id = $1", contract_id)
    if not row:
        raise HTTPException(status_code=404, detail="Договор не найден")
    return await _build_contract_response(row, pool)


@router.post("/contracts", response_model=ContractResponse, status_code=201)
async def create_contract(body: ContractCreate):
    pool = get_pool()

    # Verify farmer exists
    farmer = await pool.fetchrow("SELECT id FROM farmers WHERE id = $1", body.farmer_id)
    if not farmer:
        raise HTTPException(status_code=404, detail="Крестьянин не найден")

    # Apply tariff default if not provided
    tariff = body.tariff_amount if body.tariff_amount is not None else default_tariff(body.year)

    async with pool.acquire() as conn:
        async with conn.transaction():
            try:
                row = await conn.fetchrow("""
                    INSERT INTO contracts (farmer_id, contract_number, contract_date,
                        total_water_volume, actual_water_volume, tariff_amount, year)
                    VALUES ($1, $2, $3, $4, $5, $6, $7)
                    RETURNING id, farmer_id, contract_number, contract_date,
                        total_water_volume, actual_water_volume, tariff_amount, year,
                        created_at, updated_at
                """, body.farmer_id, body.contract_number, body.contract_date,
                    body.total_water_volume, body.actual_water_volume, tariff, body.year)
            except Exception as e:
                err = str(e)
                if "unique" in err.lower() or "duplicate" in err.lower():
                    raise HTTPException(status_code=409,
                        detail="Договор с таким номером и годом уже существует")
                raise HTTPException(status_code=500, detail=err)

            contract_id = row["id"]

            # Insert contract parcels
            for cp in body.parcels:
                geom_wkt = None
                if cp.geom:
                    geom_wkt = f"SRID=4326;{json.dumps(cp.geom)}"
                await conn.execute("""
                    INSERT INTO contract_parcels (contract_id, parcel_id, cadastral_number,
                        distribution_channel, main_channel, culture,
                        doc_hectares, irrigated_hectares, rural_district, geom)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9,
                        CASE WHEN $10::text IS NOT NULL
                             THEN ST_GeomFromGeoJSON($10)
                             ELSE NULL END)
                """, contract_id, cp.parcel_id, cp.cadastral_number,
                    cp.distribution_channel, cp.main_channel, cp.culture,
                    cp.doc_hectares, cp.irrigated_hectares, cp.rural_district,
                    json.dumps(cp.geom) if cp.geom else None)

    # Fetch farmer name for response
    farmer_row = await pool.fetchrow("SELECT full_name FROM farmers WHERE id = $1", body.farmer_id)
    row_dict = dict(row)
    row_dict["farmer_name"] = farmer_row["full_name"] if farmer_row else None
    return await _build_contract_response(row_dict, pool)


@router.put("/contracts/{contract_id}", response_model=ContractResponse)
async def update_contract(contract_id: int, body: ContractUpdate):
    pool = get_pool()
    existing = await pool.fetchrow("SELECT id FROM contracts WHERE id = $1", contract_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Договор не найден")

    fields = []
    values = []
    idx = 1
    for col in ["contract_number", "contract_date", "total_water_volume",
                "actual_water_volume", "year"]:
        val = getattr(body, col, None)
        if val is not None:
            fields.append(f"{col} = ${idx}")
            values.append(val)
            idx += 1

    # Handle tariff_amount: if explicitly set, use it. If year changed but tariff not set, apply default.
    if body.tariff_amount is not None:
        fields.append(f"tariff_amount = ${idx}")
        values.append(body.tariff_amount)
        idx += 1
    elif body.year is not None:
        dt = default_tariff(body.year)
        if dt is not None:
            fields.append(f"tariff_amount = ${idx}")
            values.append(dt)
            idx += 1

    if not fields:
        raise HTTPException(status_code=400, detail="Нет полей для обновления")

    fields.append("updated_at = NOW()")
    values.append(contract_id)

    try:
        row = await pool.fetchrow(f"""
            UPDATE contracts SET {', '.join(fields)} WHERE id = ${idx}
            RETURNING id, farmer_id, contract_number, contract_date,
                total_water_volume, actual_water_volume, tariff_amount, year,
                created_at, updated_at
        """, *values)
    except Exception as e:
        err = str(e)
        if "unique" in err.lower() or "duplicate" in err.lower():
            raise HTTPException(status_code=409,
                detail="Договор с таким номером и годом уже существует")
        raise HTTPException(status_code=500, detail=err)

    farmer_row = await pool.fetchrow("SELECT full_name FROM farmers WHERE id = $1", row["farmer_id"])
    row_dict = dict(row)
    row_dict["farmer_name"] = farmer_row["full_name"] if farmer_row else None
    return await _build_contract_response(row_dict, pool)


@router.delete("/contracts/{contract_id}")
async def delete_contract(contract_id: int):
    pool = get_pool()
    row = await pool.fetchrow("DELETE FROM contracts WHERE id = $1 RETURNING id", contract_id)
    if not row:
        raise HTTPException(status_code=404, detail="Договор не найден")
    return {"deleted": True}


# ---------- Contract Parcels management ----------

class ContractParcelUpdate(BaseModel):
    parcel_id: Optional[int] = None
    cadastral_number: Optional[str] = None
    distribution_channel: Optional[str] = None
    main_channel: Optional[str] = None
    culture: Optional[str] = None
    doc_hectares: Optional[float] = None
    irrigated_hectares: Optional[float] = None
    rural_district: Optional[str] = None
    geom: Optional[dict] = None


@router.post("/contracts/{contract_id}/parcels", response_model=ContractParcelResponse, status_code=201)
async def add_parcel_to_contract(contract_id: int, body: ContractParcelItem):
    pool = get_pool()
    existing = await pool.fetchrow("SELECT id FROM contracts WHERE id = $1", contract_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Договор не найден")

    row = await pool.fetchrow("""
        INSERT INTO contract_parcels (contract_id, parcel_id, cadastral_number,
            distribution_channel, main_channel, culture,
            doc_hectares, irrigated_hectares, rural_district, geom)
        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9,
            CASE WHEN $10::text IS NOT NULL THEN ST_GeomFromGeoJSON($10) ELSE NULL END)
        RETURNING id, contract_id, parcel_id, cadastral_number,
            distribution_channel, main_channel, culture,
            doc_hectares, irrigated_hectares, rural_district,
            ST_AsGeoJSON(geom)::jsonb AS geom_json,
            created_at, updated_at
    """, contract_id, body.parcel_id, body.cadastral_number,
        body.distribution_channel, body.main_channel, body.culture,
        body.doc_hectares, body.irrigated_hectares, body.rural_district,
        json.dumps(body.geom) if body.geom else None)

    return {
        "id": row["id"], "contract_id": row["contract_id"], "parcel_id": row["parcel_id"],
        "cadastral_number": row["cadastral_number"],
        "distribution_channel": row["distribution_channel"],
        "main_channel": row["main_channel"],
        "culture": row["culture"],
        "doc_hectares": float(row["doc_hectares"]) if row["doc_hectares"] is not None else None,
        "irrigated_hectares": float(row["irrigated_hectares"]) if row["irrigated_hectares"] is not None else None,
        "rural_district": row["rural_district"],
        "geom": json.loads(row["geom_json"]) if row["geom_json"] else None,
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }


@router.put("/contracts/{contract_id}/parcels/{cp_id}", response_model=ContractParcelResponse)
async def update_contract_parcel(contract_id: int, cp_id: int, body: ContractParcelUpdate):
    pool = get_pool()
    existing = await pool.fetchrow(
        "SELECT id FROM contract_parcels WHERE id = $1 AND contract_id = $2",
        cp_id, contract_id,
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Запись участка в договоре не найдена")

    fields = []
    values = []
    idx = 1
    for col in ["parcel_id", "cadastral_number", "distribution_channel",
                "main_channel", "culture", "doc_hectares", "irrigated_hectares",
                "rural_district"]:
        val = getattr(body, col, None)
        if val is not None:
            fields.append(f"{col} = ${idx}")
            values.append(val)
            idx += 1

    if body.geom is not None:
        fields.append(f"geom = ST_GeomFromGeoJSON(${idx})")
        values.append(json.dumps(body.geom))
        idx += 1

    if not fields:
        raise HTTPException(status_code=400, detail="Нет полей для обновления")

    fields.append("updated_at = NOW()")
    values.append(cp_id)
    idx += 1

    row = await pool.fetchrow(f"""
        UPDATE contract_parcels SET {', '.join(fields)} WHERE id = ${idx}
        RETURNING id, contract_id, parcel_id, cadastral_number,
            distribution_channel, main_channel, culture,
            doc_hectares, irrigated_hectares, rural_district,
            ST_AsGeoJSON(geom)::jsonb AS geom_json,
            created_at, updated_at
    """, *values)

    return {
        "id": row["id"], "contract_id": row["contract_id"], "parcel_id": row["parcel_id"],
        "cadastral_number": row["cadastral_number"],
        "distribution_channel": row["distribution_channel"],
        "main_channel": row["main_channel"],
        "culture": row["culture"],
        "doc_hectares": float(row["doc_hectares"]) if row["doc_hectares"] is not None else None,
        "irrigated_hectares": float(row["irrigated_hectares"]) if row["irrigated_hectares"] is not None else None,
        "rural_district": row["rural_district"],
        "geom": json.loads(row["geom_json"]) if row["geom_json"] else None,
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }


@router.delete("/contracts/{contract_id}/parcels/{cp_id}")
async def delete_contract_parcel(contract_id: int, cp_id: int):
    pool = get_pool()
    row = await pool.fetchrow(
        "DELETE FROM contract_parcels WHERE id = $1 AND contract_id = $2 RETURNING id",
        cp_id, contract_id,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Запись участка в договоре не найдена")
    return {"deleted": True}


# ---------- Excel import endpoint ----------

@router.post("/contracts/import-excel")
async def import_contracts_excel(file: UploadFile = File(...)):
    """
    Import contracts with parcels from Excel (.xlsx / .xls).
    Expected columns:
      - farmer_full_name*, farmer_iin*, farmer_phone, farmer_address
      - contract_number*, contract_date, total_water_volume, actual_water_volume, tariff_amount, year*
      - cadastral_number, distribution_channel, main_channel, culture, doc_hectares, irrigated_hectares, rural_district
    Each row = one parcel. Rows with the same contract_number+year belong to the same contract.
    Farmers are auto-created if not found by IIN.
    """
    if not file.filename or not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Поддерживаются только .xlsx / .xls файлы")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен на сервере")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    # Read header row
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip().lower().replace(" ", "_"))
        else:
            headers.append("")

    # Required columns check
    required = ["farmer_full_name", "farmer_iin", "contract_number", "year"]
    for req in required:
        if req not in headers:
            raise HTTPException(status_code=400, detail=f"В Excel отсутствует колонка: {req}")

    def col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    idx_farmer_name = col("farmer_full_name")
    idx_farmer_iin = col("farmer_iin")
    idx_farmer_phone = col("farmer_phone")
    idx_farmer_address = col("farmer_address")
    idx_contract_number = col("contract_number")
    idx_contract_date = col("contract_date")
    idx_total_water_volume = col("total_water_volume")
    idx_actual_water_volume = col("actual_water_volume")
    idx_tariff_amount = col("tariff_amount")
    idx_year = col("year")
    idx_cadastral_number = col("cadastral_number")
    idx_distribution_channel = col("distribution_channel")
    idx_main_channel = col("main_channel")
    idx_culture = col("culture")
    idx_doc_hectares = col("doc_hectares")
    idx_irrigated_hectares = col("irrigated_hectares")
    idx_rural_district = col("rural_district")

    # Group rows by (farmer_iin, contract_number, year)
    # Structure: { (iin, contract_number, year): { farmer: {}, parcels: [] } }
    groups = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        farmer_name = str(row[idx_farmer_name]).strip() if idx_farmer_name >= 0 and row[idx_farmer_name] else None
        farmer_iin = str(row[idx_farmer_iin]).strip() if idx_farmer_iin >= 0 and row[idx_farmer_iin] else None
        if not farmer_name or not farmer_iin:
            continue  # skip incomplete rows

        contract_number = str(row[idx_contract_number]).strip() if row[idx_contract_number] else None
        year_val = row[idx_year]
        try:
            year_val = int(year_val) if year_val is not None else None
        except (ValueError, TypeError):
            year_val = None

        if not contract_number or year_val is None:
            continue

        key = (farmer_iin, contract_number, year_val)

        if key not in groups:
            # Parse contract fields
            contract_date_val = None
            if idx_contract_date >= 0 and row[idx_contract_date]:
                raw_date = row[idx_contract_date]
                if hasattr(raw_date, 'date'):
                    contract_date_val = raw_date.date()
                elif isinstance(raw_date, str):
                    try:
                        from datetime import datetime as dt
                        contract_date_val = dt.strptime(raw_date, "%Y-%m-%d").date()
                    except ValueError:
                        pass

            def parse_float(v):
                if v is None:
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            groups[key] = {
                "farmer": {
                    "full_name": farmer_name,
                    "iin": farmer_iin,
                    "phone": str(row[idx_farmer_phone]).strip() if idx_farmer_phone >= 0 and row[idx_farmer_phone] else None,
                    "address": str(row[idx_farmer_address]).strip() if idx_farmer_address >= 0 and row[idx_farmer_address] else None,
                },
                "contract": {
                    "contract_number": contract_number,
                    "contract_date": contract_date_val,
                    "total_water_volume": parse_float(row[idx_total_water_volume]) if idx_total_water_volume >= 0 else None,
                    "actual_water_volume": parse_float(row[idx_actual_water_volume]) if idx_actual_water_volume >= 0 else None,
                    "tariff_amount": parse_float(row[idx_tariff_amount]) if idx_tariff_amount >= 0 else None,
                    "year": year_val,
                },
                "parcels": [],
            }

        # Add parcel
        groups[key]["parcels"].append({
            "cadastral_number": str(row[idx_cadastral_number]).strip() if idx_cadastral_number >= 0 and row[idx_cadastral_number] else None,
            "distribution_channel": str(row[idx_distribution_channel]).strip() if idx_distribution_channel >= 0 and row[idx_distribution_channel] else None,
            "main_channel": str(row[idx_main_channel]).strip() if idx_main_channel >= 0 and row[idx_main_channel] else None,
            "culture": str(row[idx_culture]).strip() if idx_culture >= 0 and row[idx_culture] else None,
            "doc_hectares": parse_float(row[idx_doc_hectares]) if idx_doc_hectares >= 0 else None,
            "irrigated_hectares": parse_float(row[idx_irrigated_hectares]) if idx_irrigated_hectares >= 0 else None,
            "rural_district": str(row[idx_rural_district]).strip() if idx_rural_district >= 0 and row[idx_rural_district] else None,
        })

    if not groups:
        raise HTTPException(status_code=400, detail="Нет данных для импорта в Excel файле")

    pool = get_pool()
    imported_contracts = 0
    imported_parcels = 0
    imported_farmers = 0
    errors = []

    async with pool.acquire() as conn:
        for (iin, cnum, year_val), group in groups.items():
            try:
                async with conn.transaction():
                    # Upsert farmer by IIN
                    farmer_row = await conn.fetchrow(
                        "SELECT id FROM farmers WHERE iin = $1", iin
                    )
                    if farmer_row:
                        farmer_id = farmer_row["id"]
                        # Update in case name/phone changed
                        await conn.execute(
                            "UPDATE farmers SET full_name=$1, phone=$2, address=$3 WHERE id=$4",
                            group["farmer"]["full_name"],
                            group["farmer"]["phone"],
                            group["farmer"]["address"],
                            farmer_id,
                        )
                    else:
                        farmer_row = await conn.fetchrow(
                            """INSERT INTO farmers (full_name, iin, phone, address)
                               VALUES ($1, $2, $3, $4) RETURNING id""",
                            group["farmer"]["full_name"],
                            group["farmer"]["iin"],
                            group["farmer"]["phone"],
                            group["farmer"]["address"],
                        )
                        farmer_id = farmer_row["id"]
                        imported_farmers += 1

                    # Set tariff default
                    tariff = group["contract"]["tariff_amount"]
                    if tariff is None:
                        tariff = default_tariff(year_val)

                    # Insert or skip contract
                    contract_row = await conn.fetchrow(
                        """INSERT INTO contracts (farmer_id, contract_number, contract_date,
                            total_water_volume, actual_water_volume, tariff_amount, year)
                           VALUES ($1, $2, $3, $4, $5, $6, $7)
                           ON CONFLICT (contract_number, year) DO UPDATE SET
                               total_water_volume = EXCLUDED.total_water_volume,
                               actual_water_volume = EXCLUDED.actual_water_volume,
                               tariff_amount = EXCLUDED.tariff_amount
                           RETURNING id, (xmax = 0) AS is_new""",
                        farmer_id,
                        group["contract"]["contract_number"],
                        group["contract"]["contract_date"],
                        group["contract"]["total_water_volume"],
                        group["contract"]["actual_water_volume"],
                        tariff,
                        group["contract"]["year"],
                    )
                    contract_id = contract_row["id"]
                    if contract_row["is_new"]:
                        imported_contracts += 1

                    # Insert parcels (delete old ones for this contract if re-importing)
                    # For simplicity, we append; use ON CONFLICT DO NOTHING is not feasible
                    # We'll just insert all
                    for p in group["parcels"]:
                        await conn.execute("""
                            INSERT INTO contract_parcels (contract_id, cadastral_number,
                                distribution_channel, main_channel, culture,
                                doc_hectares, irrigated_hectares, rural_district)
                            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                        """, contract_id,
                            p["cadastral_number"],
                            p["distribution_channel"],
                            p["main_channel"],
                            p["culture"],
                            p["doc_hectares"],
                            p["irrigated_hectares"],
                            p["rural_district"])
                        imported_parcels += 1

            except Exception as e:
                errors.append(f"Строка ({group['farmer']['full_name']}, {cnum}, {year_val}): {e}")

    return {
        "imported": {
            "farmers": imported_farmers,
            "contracts": imported_contracts,
            "parcels": imported_parcels,
        },
        "errors": errors,
    }
