import io
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from fastapi import HTTPException
from ..schemas.farmer import FarmerCreate, FarmerUpdate


def _row_to_dict(row) -> dict:
    return {
        "id": row["id"], "full_name": row["full_name"], "iin": row["iin"],
        "contract_number": row["contract_number"], "phone": row["phone"],
        "address": row["address"],
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }


async def list_farmers(db: AsyncSession, search: Optional[str] = None) -> list:
    if search:
        term = f"%{search.strip()}%"
        result = await db.execute(text("""
            SELECT id, full_name, iin, contract_number, phone, address, created_at, updated_at
            FROM farmers
            WHERE full_name ILIKE :t OR iin ILIKE :t OR contract_number ILIKE :t
            ORDER BY id
        """), {"t": term})
    else:
        result = await db.execute(text("""
            SELECT id, full_name, iin, contract_number, phone, address, created_at, updated_at
            FROM farmers ORDER BY id
        """))
    return [_row_to_dict(r) for r in result.mappings().all()]


async def get_farmer(db: AsyncSession, farmer_id: int) -> dict:
    result = await db.execute(text("""
        SELECT id, full_name, iin, contract_number, phone, address, created_at, updated_at
        FROM farmers WHERE id = :id
    """), {"id": farmer_id})
    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Крестьянин не найден")
    return _row_to_dict(row)


async def create_farmer(db: AsyncSession, body: FarmerCreate) -> dict:
    try:
        result = await db.execute(text("""
            INSERT INTO farmers (full_name, iin, contract_number, phone, address)
            VALUES (:full_name, :iin, :contract_number, :phone, :address)
            RETURNING id, full_name, iin, contract_number, phone, address, created_at, updated_at
        """), body.model_dump())
        await db.commit()
        return _row_to_dict(result.mappings().first())
    except Exception as e:
        await db.rollback()
        err = str(e)
        if "unique" in err.lower() or "duplicate" in err.lower():
            raise HTTPException(status_code=409, detail="Крестьянин с таким ИИН уже существует")
        raise HTTPException(status_code=500, detail=err)


async def update_farmer(db: AsyncSession, farmer_id: int, body: FarmerUpdate) -> dict:
    existing = await db.execute(text("SELECT id FROM farmers WHERE id = :id"), {"id": farmer_id})
    if not existing.mappings().first():
        raise HTTPException(status_code=404, detail="Крестьянин не найден")

    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Нет полей для обновления")

    set_clause = ", ".join(f"{k} = :{k}" for k in updates)
    updates["id"] = farmer_id

    try:
        result = await db.execute(text(f"""
            UPDATE farmers SET {set_clause}, updated_at = NOW()
            WHERE id = :id
            RETURNING id, full_name, iin, contract_number, phone, address, created_at, updated_at
        """), updates)
        await db.commit()
        return _row_to_dict(result.mappings().first())
    except Exception as e:
        await db.rollback()
        err = str(e)
        if "unique" in err.lower() or "duplicate" in err.lower():
            raise HTTPException(status_code=409, detail="Крестьянин с таким ИИН уже существует")
        raise HTTPException(status_code=500, detail=err)


async def delete_farmer(db: AsyncSession, farmer_id: int) -> dict:
    result = await db.execute(text(
        "DELETE FROM farmers WHERE id = :id RETURNING id"
    ), {"id": farmer_id})
    await db.commit()
    if not result.mappings().first():
        raise HTTPException(status_code=404, detail="Крестьянин не найден")
    return {"deleted": True}


async def import_excel(db: AsyncSession, file_bytes: bytes) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower().replace(" ", "_") if c.value else "" for c in ws[1]]
    for req in ["full_name", "iin"]:
        if req not in headers:
            raise HTTPException(status_code=400, detail=f"Отсутствует колонка: {req}")

    def ci(name): return headers.index(name) if name in headers else -1

    imported, errors = 0, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def cell(n): return str(row[ci(n)]).strip() if ci(n) >= 0 and row[ci(n)] else None
        full_name = cell("full_name")
        iin = cell("iin")
        if not full_name or not iin:
            continue
        try:
            existing = await db.execute(text("SELECT id FROM farmers WHERE iin = :iin"), {"iin": iin})
            ex = existing.mappings().first()
            if ex:
                await db.execute(text("""
                    UPDATE farmers SET full_name=:fn, contract_number=:cn, phone=:ph, address=:ad
                    WHERE id=:id
                """), {"fn": full_name, "cn": cell("contract_number"), "ph": cell("phone"),
                       "ad": cell("address"), "id": ex["id"]})
            else:
                await db.execute(text("""
                    INSERT INTO farmers (full_name, iin, contract_number, phone, address)
                    VALUES (:fn, :iin, :cn, :ph, :ad)
                """), {"fn": full_name, "iin": iin, "cn": cell("contract_number"),
                       "ph": cell("phone"), "ad": cell("address")})
                imported += 1
            await db.commit()
        except Exception as e:
            await db.rollback()
            errors.append(f"({full_name}, {iin}): {e}")

    return {"imported": imported, "errors": errors}
