import io
import json
from typing import Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from fastapi import HTTPException
from ..schemas.contract import ContractCreate, ContractUpdate, ContractParcelItem

TARIFF_BY_YEAR = {
    2019: 1.0, 2020: 1.0, 2021: 1.0, 2022: 1.0, 2023: 1.0, 2024: 1.0,
    2026: 1.8,
}

CONTRACT_WITH_FARMER = """
    SELECT c.id, c.farmer_id, f.full_name AS farmer_name,
           c.contract_number, c.contract_date, c.total_water_volume,
           c.actual_water_volume, c.tariff_amount,
           c.year, c.created_at, c.updated_at
    FROM contracts c
    LEFT JOIN farmers f ON f.id = c.farmer_id
"""


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

async def _fetch_parcels(db: AsyncSession, contract_id: int) -> list:
    result = await db.execute(text("""
        SELECT cp.id, cp.contract_id, cp.parcel_id, cp.cadastral_number,
               cp.distribution_channel, cp.main_channel, cp.culture,
               cp.doc_hectares, cp.irrigated_hectares, cp.rural_district,
               ST_AsGeoJSON(COALESCE(cp.geom, ST_Transform(p.geom, 4326)))::jsonb AS geom_json,
               cp.created_at, cp.updated_at
        FROM contract_parcels cp
        LEFT JOIN parcels p ON p.id = cp.parcel_id
        WHERE cp.contract_id = :cid ORDER BY cp.id
    """), {"cid": contract_id})
    return [
        {
            "id": r["id"], "contract_id": r["contract_id"], "parcel_id": r["parcel_id"],
            "cadastral_number": r["cadastral_number"],
            "distribution_channel": r["distribution_channel"],
            "main_channel": r["main_channel"], "culture": r["culture"],
            "doc_hectares": float(r["doc_hectares"]) if r["doc_hectares"] is not None else None,
            "irrigated_hectares": float(r["irrigated_hectares"]) if r["irrigated_hectares"] is not None else None,
            "rural_district": r["rural_district"],
            "geom": r["geom_json"],
            "created_at": r["created_at"].isoformat(),
            "updated_at": r["updated_at"].isoformat(),
        }
        for r in result.mappings().all()
    ]


def _contract_dict(row, parcels: list) -> dict:
    return {
        "id": row["id"], "farmer_id": row["farmer_id"],
        "farmer_name": row["farmer_name"],
        "contract_number": row["contract_number"],
        "contract_date": row["contract_date"].isoformat() if row["contract_date"] else None,
        "total_water_volume": float(row["total_water_volume"]) if row["total_water_volume"] is not None else None,
        "actual_water_volume": float(row["actual_water_volume"]) if row["actual_water_volume"] is not None else None,
        "tariff_amount": float(row["tariff_amount"]) if row["tariff_amount"] is not None else None,
        "year": row["year"],
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
        "parcels": parcels,
    }


async def _resolve_parcel_id(db: AsyncSession, cadastral_number: Optional[str]) -> Optional[int]:
    """Ищет parcel_id по cadastral_number в таблице parcels."""
    if not cadastral_number:
        return None
    result = await db.execute(
        text("SELECT id FROM parcels WHERE cadastral_number = :cn LIMIT 1"),
        {"cn": cadastral_number.strip()}
    )
    row = result.mappings().first()
    return row["id"] if row else None


async def _insert_contract_parcel(db: AsyncSession, contract_id: int, p: ContractParcelItem):
    """Вставляет один участок договора, автоматически подтягивая parcel_id."""
    parcel_id = p.parcel_id
    if parcel_id is None:
        parcel_id = await _resolve_parcel_id(db, p.cadastral_number)

    geom_val = f"ST_SetSRID(ST_GeomFromGeoJSON('{json.dumps(p.geom)}'),4326)" if p.geom else "NULL"
    await db.execute(text(f"""
        INSERT INTO contract_parcels
            (contract_id, parcel_id, cadastral_number, distribution_channel,
             main_channel, culture, doc_hectares, irrigated_hectares, rural_district, geom)
        VALUES
            (:cid, :pid, :cn, :dc, :mc, :cu, :dh, :ih, :rd, {geom_val})
    """), {
        "cid": contract_id, "pid": parcel_id,
        "cn": p.cadastral_number, "dc": p.distribution_channel,
        "mc": p.main_channel, "cu": p.culture,
        "dh": p.doc_hectares, "ih": p.irrigated_hectares, "rd": p.rural_district,
    })


# ─────────────────────────────────────────────
# CRUD
# ─────────────────────────────────────────────

async def list_contracts(
    db: AsyncSession,
    farmer_id: Optional[int] = None,
    year: Optional[int] = None,
    search: Optional[str] = None,
) -> list:
    conditions, params = [], {}
    if farmer_id is not None:
        conditions.append("c.farmer_id = :farmer_id"); params["farmer_id"] = farmer_id
    if year is not None:
        conditions.append("c.year = :year"); params["year"] = year
    if search:
        conditions.append("c.contract_number ILIKE :search")
        params["search"] = f"%{search.strip()}%"
    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    result = await db.execute(text(CONTRACT_WITH_FARMER + where + " ORDER BY c.year DESC, c.id"), params)
    out = []
    for r in result.mappings().all():
        parcels = await _fetch_parcels(db, r["id"])
        out.append(_contract_dict(r, parcels))
    return out


async def get_contract(db: AsyncSession, contract_id: int) -> dict:
    result = await db.execute(
        text(CONTRACT_WITH_FARMER + " WHERE c.id = :id"), {"id": contract_id}
    )
    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Договор не найден")
    return _contract_dict(row, await _fetch_parcels(db, contract_id))


async def create_contract(db: AsyncSession, body: ContractCreate) -> dict:
    try:
        result = await db.execute(text("""
            INSERT INTO contracts
                (farmer_id, contract_number, contract_date, total_water_volume,
                 actual_water_volume, tariff_amount, year)
            VALUES
                (:farmer_id, :contract_number, :contract_date, :total_water_volume,
                 :actual_water_volume, :tariff_amount, :year)
            RETURNING id
        """), {
            "farmer_id": body.farmer_id, "contract_number": body.contract_number,
            "contract_date": body.contract_date, "total_water_volume": body.total_water_volume,
            "actual_water_volume": body.actual_water_volume,
            "tariff_amount": body.tariff_amount, "year": body.year,
        })
        contract_id = result.mappings().first()["id"]
        for p in body.parcels:
            await _insert_contract_parcel(db, contract_id, p)
        await db.commit()
        return await get_contract(db, contract_id)
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        err = str(e)
        if "unique" in err.lower():
            raise HTTPException(status_code=409, detail="Договор с таким номером и годом уже существует")
        raise HTTPException(status_code=500, detail=err)


async def update_contract(db: AsyncSession, contract_id: int, body: ContractUpdate) -> dict:
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Нет полей для обновления")
    set_clause = ", ".join(f"{k} = :{k}" for k in updates)
    updates["id"] = contract_id
    try:
        result = await db.execute(
            text(f"UPDATE contracts SET {set_clause}, updated_at=NOW() WHERE id=:id RETURNING id"),
            updates
        )
        if not result.mappings().first():
            raise HTTPException(status_code=404, detail="Договор не найден")
        await db.commit()
        return await get_contract(db, contract_id)
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


async def update_contract_with_parcels(
    db: AsyncSession,
    contract_id: int,
    body: ContractUpdate,
    parcels: List[ContractParcelItem],
) -> dict:
    """Обновляет договор + полностью заменяет его участки."""
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if updates:
        set_clause = ", ".join(f"{k} = :{k}" for k in updates)
        updates["id"] = contract_id
        result = await db.execute(
            text(f"UPDATE contracts SET {set_clause}, updated_at=NOW() WHERE id=:id RETURNING id"),
            updates
        )
        if not result.mappings().first():
            await db.rollback()
            raise HTTPException(status_code=404, detail="Договор не найден")

    # Удаляем старые участки и вставляем новые
    await db.execute(
        text("DELETE FROM contract_parcels WHERE contract_id = :cid"),
        {"cid": contract_id}
    )
    for p in parcels:
        await _insert_contract_parcel(db, contract_id, p)

    try:
        await db.commit()
        return await get_contract(db, contract_id)
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


async def delete_contract(db: AsyncSession, contract_id: int) -> dict:
    result = await db.execute(
        text("DELETE FROM contracts WHERE id=:id RETURNING id"), {"id": contract_id}
    )
    await db.commit()
    if not result.mappings().first():
        raise HTTPException(status_code=404, detail="Договор не найден")
    return {"deleted": True}


# ─────────────────────────────────────────────
# БАГ 1 — GET /api/contracts/parcels/map
# ─────────────────────────────────────────────

async def get_contracts_map(db: AsyncSession) -> dict:
    """
    GeoJSON для карты: контрактные участки с геометрией из contract_parcels.
    Возвращает outlines + fills в том же формате что и /api/parcels/map.
    """
    result = await db.execute(text("""
        SELECT
            cp.id, cp.cadastral_number, cp.culture, cp.rural_district,
            cp.doc_hectares, cp.irrigated_hectares,
            c.contract_number, c.year,
            f.full_name AS farmer_name,
            COALESCE(cp.geom, ST_Transform(p.geom, 4326)) AS geom
        FROM contract_parcels cp
        JOIN contracts c ON c.id = cp.contract_id
        LEFT JOIN farmers f ON f.id = c.farmer_id
        LEFT JOIN parcels p ON p.id = cp.parcel_id
        WHERE COALESCE(cp.geom, ST_Transform(p.geom, 4326)) IS NOT NULL
    """))
    rows = result.mappings().all()

    outlines, fills = [], []
    for row in rows:
        props = {
            "id": row["id"],
            "cadastral_number": row["cadastral_number"],
            "culture": row["culture"],
            "rural_district": row["rural_district"],
            "doc_hectares": float(row["doc_hectares"]) if row["doc_hectares"] else None,
            "irrigated_hectares": float(row["irrigated_hectares"]) if row["irrigated_hectares"] else None,
            "contract_number": row["contract_number"],
            "year": row["year"],
            "farmer_name": row["farmer_name"],
            "fill_color": [88, 86, 214, 140],   # фиолетовый для договорных участков
            "fill_pct": 100,
            "status": "contract",
        }
        geom_json = json.loads(row["geom"]) if isinstance(row["geom"], str) else row["geom"]
        feature = {"type": "Feature", "geometry": geom_json, "properties": props}
        outlines.append(feature)
        fills.append(feature)

    return {
        "outlines": {"type": "FeatureCollection", "features": outlines},
        "fills":    {"type": "FeatureCollection", "features": fills},
    }


# ─────────────────────────────────────────────
# БАГ 2 — POST /api/contracts/import-excel
# ─────────────────────────────────────────────

async def import_excel(db: AsyncSession, file_bytes: bytes) -> dict:
    """
    Импорт договоров из Excel.
    Ожидаемые колонки:
      contract_number*, year*, farmer_iin или farmer_name*,
      contract_date, total_water_volume, actual_water_volume, tariff_amount,
      cadastral_number, distribution_channel, main_channel, culture,
      doc_hectares, irrigated_hectares, rural_district
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    raw_headers = [str(c.value).strip().lower().replace(" ", "_") if c.value else "" for c in ws[1]]

    required = ["contract_number"]
    for req in required:
        if req not in raw_headers:
            raise HTTPException(status_code=400, detail=f"Отсутствует обязательная колонка: {req}")

    def ci(name): return raw_headers.index(name) if name in raw_headers else -1
    def cell(row, name):
        idx = ci(name)
        return str(row[idx]).strip() if idx >= 0 and row[idx] is not None else None
    def cell_float(row, name):
        v = cell(row, name)
        try: return float(v) if v else None
        except: return None
    def cell_int(row, name):
        v = cell(row, name)
        try: return int(float(v)) if v else None
        except: return None

    imported = {"farmers": 0, "contracts": 0, "parcels": 0}
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        contract_number = cell(row, "contract_number")
        if not contract_number:
            continue

        try:
            # ── 1. Найти или создать крестьянина ──
            farmer_iin  = cell(row, "farmer_iin") or cell(row, "iin")
            farmer_name = cell(row, "farmer_name") or cell(row, "full_name")
            farmer_id   = None

            if farmer_iin:
                r = (await db.execute(
                    text("SELECT id FROM farmers WHERE iin = :iin"), {"iin": farmer_iin}
                )).mappings().first()
                if r:
                    farmer_id = r["id"]
                elif farmer_name:
                    r = (await db.execute(text("""
                        INSERT INTO farmers (full_name, iin) VALUES (:fn, :iin)
                        RETURNING id
                    """), {"fn": farmer_name, "iin": farmer_iin})).mappings().first()
                    farmer_id = r["id"]
                    imported["farmers"] += 1
            elif farmer_name:
                r = (await db.execute(
                    text("SELECT id FROM farmers WHERE full_name = :fn LIMIT 1"),
                    {"fn": farmer_name}
                )).mappings().first()
                farmer_id = r["id"] if r else None

            if not farmer_id:
                errors.append(f"Строка {contract_number}: крестьянин не найден и не создан")
                continue

            # ── 2. Найти или создать договор ──
            year = cell_int(row, "year")
            existing_c = (await db.execute(text("""
                SELECT id FROM contracts
                WHERE contract_number = :cn AND (year = :yr OR (:yr IS NULL AND year IS NULL))
            """), {"cn": contract_number, "yr": year})).mappings().first()

            if existing_c:
                contract_id = existing_c["id"]
            else:
                r = (await db.execute(text("""
                    INSERT INTO contracts
                        (farmer_id, contract_number, contract_date, total_water_volume,
                         actual_water_volume, tariff_amount, year)
                    VALUES (:fid, :cn, :cd, :twv, :awv, :ta, :yr)
                    RETURNING id
                """), {
                    "fid": farmer_id, "cn": contract_number,
                    "cd": cell(row, "contract_date"),
                    "twv": cell_float(row, "total_water_volume"),
                    "awv": cell_float(row, "actual_water_volume"),
                    "ta": cell_float(row, "tariff_amount") or TARIFF_BY_YEAR.get(year),
                    "yr": year,
                })).mappings().first()
                contract_id = r["id"]
                imported["contracts"] += 1

            # ── 3. Участок договора ──
            cadastral = cell(row, "cadastral_number")
            if cadastral:
                # БАГ 3 FIX: ищем parcel_id по cadastral_number
                parcel_id = await _resolve_parcel_id(db, cadastral)

                await db.execute(text("""
                    INSERT INTO contract_parcels
                        (contract_id, parcel_id, cadastral_number, distribution_channel,
                         main_channel, culture, doc_hectares, irrigated_hectares, rural_district)
                    VALUES
                        (:cid, :pid, :cn, :dc, :mc, :cu, :dh, :ih, :rd)
                    ON CONFLICT DO NOTHING
                """), {
                    "cid": contract_id,
                    "pid": parcel_id,           # ← теперь не NULL если есть в parcels
                    "cn":  cadastral,
                    "dc":  cell(row, "distribution_channel"),
                    "mc":  cell(row, "main_channel"),
                    "cu":  cell(row, "culture"),
                    "dh":  cell_float(row, "doc_hectares"),
                    "ih":  cell_float(row, "irrigated_hectares"),
                    "rd":  cell(row, "rural_district"),
                })
                imported["parcels"] += 1

            await db.commit()

        except Exception as e:
            await db.rollback()
            errors.append(f"Строка ({contract_number}): {e}")

    return {"imported": imported, "errors": errors}


# ─────────────────────────────────────────────
# Stats
# ─────────────────────────────────────────────

async def get_contracts_stats(
    db: AsyncSession,
    year: Optional[int] = None,
    rural_district: Optional[str] = None,
    culture: Optional[str] = None,
    farmer_id: Optional[int] = None,
) -> dict:
    c_conds, p_conds, params = [], [], {}
    if year is not None:
        c_conds.append("c.year = :year"); p_conds.append("c.year = :year"); params["year"] = year
    if farmer_id is not None:
        c_conds.append("c.farmer_id = :farmer_id"); p_conds.append("c.farmer_id = :farmer_id")
        params["farmer_id"] = farmer_id
    if rural_district:
        p_conds.append("cp.rural_district ILIKE :rd"); params["rd"] = f"%{rural_district.strip()}%"
    if culture:
        p_conds.append("cp.culture ILIKE :culture"); params["culture"] = f"%{culture.strip()}%"

    c_where = (" WHERE " + " AND ".join(c_conds)) if c_conds else ""
    p_where = (" WHERE " + " AND ".join(p_conds)) if p_conds else ""

    c_row = (await db.execute(text(f"""
        SELECT COUNT(*) AS total_contracts,
               COALESCE(SUM(c.total_water_volume),0) AS total_plan_volume,
               COALESCE(SUM(c.actual_water_volume),0) AS total_actual_volume,
               COALESCE(SUM(c.tariff_amount),0) AS total_tariff
        FROM contracts c {c_where}
    """), params)).mappings().first()

    p_row = (await db.execute(text(f"""
        SELECT COUNT(*) AS total_parcels,
               COALESCE(SUM(cp.doc_hectares),0) AS total_doc_hectares,
               COALESCE(SUM(cp.irrigated_hectares),0) AS total_irrigated_hectares
        FROM contract_parcels cp JOIN contracts c ON c.id=cp.contract_id {p_where}
    """), params)).mappings().first()

    years = (await db.execute(text(f"""
        SELECT c.year, COUNT(*) AS cnt,
               COALESCE(SUM(c.total_water_volume),0) AS plan_vol,
               COALESCE(SUM(c.actual_water_volume),0) AS actual_vol,
               COALESCE(SUM(c.tariff_amount),0) AS tariff
        FROM contracts c {c_where} GROUP BY c.year ORDER BY c.year
    """), params)).mappings().all()

    cultures = (await db.execute(text(f"""
        SELECT cp.culture, COUNT(*) AS cnt,
               COALESCE(SUM(cp.doc_hectares),0) AS doc_ha,
               COALESCE(SUM(cp.irrigated_hectares),0) AS irr_ha
        FROM contract_parcels cp JOIN contracts c ON c.id=cp.contract_id {p_where}
        GROUP BY cp.culture ORDER BY cnt DESC
    """), params)).mappings().all()

    return {
        "overview": {
            "total_contracts": c_row["total_contracts"],
            "total_plan_volume": float(c_row["total_plan_volume"]),
            "total_actual_volume": float(c_row["total_actual_volume"]),
            "total_tariff": float(c_row["total_tariff"]),
            "total_parcels": p_row["total_parcels"],
            "total_doc_hectares": float(p_row["total_doc_hectares"]),
            "total_irrigated_hectares": float(p_row["total_irrigated_hectares"]),
        },
        "by_year": [
            {"year": r["year"], "contracts": r["cnt"],
             "plan_volume": float(r["plan_vol"]),
             "actual_volume": float(r["actual_vol"]),
             "tariff": float(r["tariff"])}
            for r in years
        ],
        "by_culture": [
            {"culture": r["culture"], "parcels": r["cnt"],
             "doc_hectares": float(r["doc_ha"]),
             "irrigated_hectares": float(r["irr_ha"])}
            for r in cultures
        ],
    }
